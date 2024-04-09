#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2024/3/23 11:44
# @Author  : Lihua
# @Version : 1.0
# @Contact : heei3k@hotmail.com
# @File    : dianping.py
# @Software: PyCharm

import requests
from bs4 import BeautifulSoup
from openpyxl.workbook import Workbook


def dianping_spider(keywords):
    wb = Workbook()
    ws = wb.active

    ws['A1'] = '店铺名称'
    ws['B1'] = "评论数"
    ws['C1'] = "人均消费"
    ws['D1'] = "类型标签"
    ws['E1'] = "所在地区"
    ws['F1'] = "推荐菜"
    ws['G1'] = "团购信息"
    ws['H1'] = "详细地址"
    ws['I1'] = "联系电话"
    ws['J1'] = "评论打分"
    ws['K1'] = "首条评论"
    ws['L1'] = "第二条评论"
    ws['M1'] = "第三条评论"
    ws['N1'] = "第四条评论"
    ws['O1'] = "第五条评论"

    cookies = {
        'fspop': 'test',
        'cy': '5',
        'cye': 'nanjing',
        '_lxsdk_cuid': '188469ba842c8-0516f336e1b5048-d565429-144000-188469ba842c8',
        '_lxsdk': '188469ba842c8-0516f336e1b5048-d565429-144000-188469ba842c8',
        '_hc.v': '94fffe21-dda9-b0b8-f9ed-9b9831c64cb9.1711164134',
        'WEBDFPID': '9140y9xw93yv5820y3802w4y4xzuxwyw81v3045w73297958vx321zz0-2026524150568-1711164150568AWEQQMQ10f02007e9804b0b4cf483cebf1f9f511971',
        'Hm_lvt_602b80cf8079ae6591966cc70a3940e7': '1711164135,1711266450',
        's_ViewType': '10',
        'qruuid': '5901f898-5f77-4db0-9033-f9b2e5b82cfb',
        'ctu': '86597371eb181d9afe497f3eecc833aca8e59a82e7b5369bab615a67d8eb3bf8',
        'dplet': 'ca5ac86bb96011c322dba6e8909cbde9',
        'dper': '02021325a59275e40fc84a500c1ab72583d130dda3873754f8323a1f05dce962428a89806d4790b3498bbec01014c264ff71da07b5ba7b53dac300000000fa1e000044fe9d897bd5f54eb62fdeb40623aa14e7fc40c61a6676887cf6f4a7ce53b51cc54d270d4fee74b458aa4ca64b0347fd',
        'ua': '^%^E8^%^B4^%^9D^%^E8^%^B4^%^9D^%^E4^%^B9^%^96^%^E4^%^B9^%^962016',
        'll': '7fd06e815b796be3df069dec7836c3df',
        '_lx_utm': 'utm_source^%^3DBaidu^%^26utm_medium^%^3Dorganic',
        '_lxsdk_s': '18e6f6dc818-b4c-38c-7cc^%^7C^%^7C248',
        'Hm_lpvt_602b80cf8079ae6591966cc70a3940e7': '1711268555',
    }

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
        'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
        # 'Accept-Encoding': 'gzip, deflate, br',
        'Referer': 'https://account.dianping.com/',
        'Connection': 'keep-alive',
        # 'Cookie': 'fspop=test; cy=5; cye=nanjing; _lxsdk_cuid=188469ba842c8-0516f336e1b5048-d565429-144000-188469ba842c8; _lxsdk_s=18e69549367-e14-04d-e72^%^7C^%^7C88; _lxsdk=188469ba842c8-0516f336e1b5048-d565429-144000-188469ba842c8; _hc.v=94fffe21-dda9-b0b8-f9ed-9b9831c64cb9.1711164134; WEBDFPID=9140y9xw93yv5820y3802w4y4xzuxwyw81v3045w73297958vx321zz0-2026524150568-1711164150568AWEQQMQ10f02007e9804b0b4cf483cebf1f9f511971; Hm_lvt_602b80cf8079ae6591966cc70a3940e7=1711164135; Hm_lpvt_602b80cf8079ae6591966cc70a3940e7=1711164400; s_ViewType=10; qruuid=42dbc9e4-dd94-4904-9479-0c5ab28c86a5; dplet=6f33d38e27b73460484774fdf871f300; dper=02026526ceae3aa351d563ee49c328f6c7c4d77968526ec7baa4205c90869514bc46fb829059e0fee8b1be91b06947180e9c6dd8ef8e56402d7400000000fa1e000066f5b188cee1ab7a2c19776763a742c6e033a17519b899bd13f6b205d6eaf72795d1abccc5cda1dc5d0ced0747bd672c; ll=7fd06e815b796be3df069dec7836c3df; ua=^%^E8^%^B4^%^9D^%^E8^%^B4^%^9D^%^E4^%^B9^%^96^%^E4^%^B9^%^962016; ctu=86597371eb181d9afe497f3eecc833aca8e59a82e7b5369bab615a67d8eb3bf8',
        'Upgrade-Insecure-Requests': '1',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'same-site',
        'Sec-Fetch-User': '?1',
        'Pragma': 'no-cache',
        'Cache-Control': 'no-cache',
    }

    for i in range(1, 51):
        if i == 1:
            if keywords is None:
                url = 'https://www.dianping.com/nanjing/ch10/g110'
            else:
                url = 'https://www.dianping.com/search/keyword/5/0_' + keywords
        else:
            if keywords is None:
                url = f'https://www.dianping.com/nanjing/ch10/g110p{str(i)}'
            else:
                url = f'https://www.dianping.com/search/keyword/5/0_{keywords}/p{str(i)}'

        response = requests.get(url, cookies=cookies, headers=headers)

        soup = BeautifulSoup(response.text, "lxml")

        soup.find()

        shop_list = soup.find_all("div", id="shop-all-list")[0].find_all("li")

        for shop in shop_list:
            name = shop.find_all("h4")[0].text.strip()
            shop_href = shop.find_all("div", class_="tit")[0].find_all("a")[0].get("href")
            try:
                # data-click-name不是一个标准的html标签，所以不能直接查找，可以通过lambda表达式来筛选出需要的标签
                review = shop.find_all("a", attrs={'data-click-name': 'shop_iwant_review_click'})[0].find_all("b")[
                    0].text.strip()
            except IndexError as e:
                review = '无评论'
                print(name)
            try:
                mean_price = shop.find_all("a", class_="mean-price")[0].find_all("b")[0].text.strip()
            except:
                mean_price = '-'
                print(name)
            tag = shop.find_all("div", class_="tag-addr")[0].find_all("a")[0].text.strip()
            region = shop.find_all("div", class_="tag-addr")[0].find_all("a")[1].text.strip()
            recommend = ' '.join(i.text.strip() for i in shop.find_all("a", class_="recommend-click"))
            try:
                svr_info = shop.find_all("div", class_="svr-info")[0].find_all("a", target="_blank")[0].get(
                    "title").strip()
            except:
                svr_info = '无'
            print(name, review, mean_price, tag, region, recommend, svr_info)
            ws.append([name, review, mean_price, tag, region, recommend, svr_info])

            # res= requests.get(shop_href, cookies=cookies, headers=headers)
            # soup = BeautifulSoup(res.text, "lxml")
            # try:
            #     svr_info = shop.find_all("div", class_="svr-info")[0].find_all("a", target="_blank")[0].get("title").strip()
            #     address=soup.find_all("span", id="address")[0].text.strip()
            #     telephone=soup.find_all("p", class_="expand-info tel")[0].text.strip()
            # except IndexError:
            #     continue
            # if svr_info and address and telephone:
            #     ws.append([name,review, mean_price, tag, region, recommend, svr_info,address,telephone])
            #     # print(name, review, mean_price, tag, region, recommend, svr_info, address, telephone)
            # elif svr_info:
            #     ws.append([name, review, mean_price, tag, region, recommend, svr_info])
            # else:
            #     ws.append([name,review, mean_price, tag, region, recommend])
            #     # print(name, review, mean_price, tag, region, recommend, svr_info)
    if keywords is None:
        keywords = "boss"
    wb.save(f'D:/Download/export_data/dianping/{keywords}.xlsx')


# print(shop_list)
if __name__ == '__main__':
    dianping_spider('烧烤')
