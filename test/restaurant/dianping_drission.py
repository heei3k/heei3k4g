#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2024/3/23 11:44
# @Author  : Lihua
# @Version : 1.0
# @Contact : heei3k@hotmail.com
# @File    : dianping.py
# @Software: PyCharm
from DrissionPage import ChromiumPage
from openpyxl.workbook import Workbook


def dianping_spider(keywords):
    wb = Workbook()
    ws = wb.active
    page = ChromiumPage()

    ws['A1'] = '店铺名称'
    ws['B1'] = "评论数"
    ws['C1'] = "人均消费"
    ws['D1'] = "类型标签"
    ws['E1'] = "所在地区"
    ws['F1'] = "推荐菜"
    ws['G1'] = "团购信息"

    if keywords is None:
        url = 'https://www.dianping.com/nanjing/ch10/g110'
    else:
        url = 'https://www.dianping.com/search/keyword/5/0_' + keywords

    page.get(url)
    print(page.response)
    print(page.response.text)

    # for i in range(1,51):
    #     response =page.get(url)
    #     print(response.status_code)

    # soup = BeautifulSoup(response.text, "lxml")
    #
    # soup.find()
    #
    # shop_list = soup.find_all("div", id="shop-all-list")[0].find_all("li")
    #
    # for shop in shop_list:
    #     name = shop.find_all("h4")[0].text.strip()
    #     shop_href = shop.find_all("div", class_="tit")[0].find_all("a")[0].get("href")
    #     try:
    #         # data-click-name不是一个标准的html标签，所以不能直接查找，可以通过lambda表达式来筛选出需要的标签
    #         review = shop.find_all("a",attrs={'data-click-name':'shop_iwant_review_click'})[0].find_all("b")[0].text.strip()
    #     except IndexError as e:
    #         review = '无评论'
    #         print(name)
    #     try:
    #         mean_price = shop.find_all("a", class_="mean-price")[0].find_all("b")[0].text.strip()
    #     except:
    #         mean_price = '-'
    #         print(name)
    #     tag = shop.find_all("div", class_="tag-addr")[0].find_all("a")[0].text.strip()
    #     region= shop.find_all("div", class_="tag-addr")[0].find_all("a")[1].text.strip()
    #     recommend = ' '.join(i.text.strip() for i in shop.find_all("a", class_="recommend-click"))
    #     try:
    #         svr_info = shop.find_all("div", class_="svr-info")[0].find_all("a", target="_blank")[0].get("title").strip()
    #     except:
    #         svr_info = '无'
    #     print(name, review, mean_price, tag, region, recommend,svr_info)
    #     ws.append([name, review, mean_price, tag, region, recommend, svr_info])

    # wb.save('D:/Download/export_data/dianping/dianping.xlsx')


if __name__ == '__main__':
    ds = dianping_spider('汤包')
