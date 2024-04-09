#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2024/3/23 17:53
# @Author  : Lihua
# @Version : 1.0
# @Contact : heei3k@hotmail.com
# @File    : boss.py
# @Software: PyCharm
from DrissionPage import ChromiumPage
from openpyxl.workbook import Workbook


def boss_spider(keywords=None):
    wb = Workbook()
    ws = wb.active

    if keywords is None:
        url = f"https://www.zhipin.com/web/geek/job?city=101190100"
        keywords = 'boss'
    else:
        url = f"https://www.zhipin.com/web/geek/job?query={keywords}&city=101190100"

    ws['A1'] = '公司名'
    ws['B1'] = "行业"
    ws['C1'] = "人员规模"
    ws['D1'] = "联系人"
    ws['E1'] = "岗位名称"
    ws['F1'] = "薪资范围"
    ws['G1'] = "年限及学历要求"
    ws['H1'] = "要求技能"
    ws['I1'] = "工作城市"
    ws['J1'] = "所在地区"
    ws['K1'] = "所在商圈"
    ws['L1'] = "福利待遇"

    driver = ChromiumPage()
    driver.listen.start('/wapi/zpgeek/search/joblist.json')
    driver.get(url)

    for i in range(10):
        if i > 0:
            driver('@class=ui-icon-arrow-right').click()
        resp = driver.listen.wait(timeout=10)
        if resp is not False:
            json_response = resp.response.body
        # print(json_response)
        if json_response:
            zpData = json_response['zpData']
        else:
            zpData = None
        if zpData and 'jobList' in zpData:
            jobList = zpData['jobList']

            for job in jobList:
                brandName = job['brandName']
                brandIndustry = job['brandIndustry']
                brandScaleName = job['brandScaleName']
                bossName = job['bossName']
                jobName = job['jobName']
                salaryDesc = job['salaryDesc']
                jobLabels = ','.join(job['jobLabels'])
                skills = ','.join(job['skills'])
                cityName = job['cityName']
                areaDistrict = job['areaDistrict']
                businessDistrict = job['businessDistrict']
                welfareList = ','.join(job['welfareList'])
                ws.append([brandName, brandIndustry, brandScaleName, bossName, jobName, salaryDesc, jobLabels, skills,
                           cityName, areaDistrict, businessDistrict, welfareList])
                print(
                    f'{brandName} {brandIndustry} {brandScaleName} {bossName} {jobName} {salaryDesc} {jobLabels} {skills} {cityName} {areaDistrict} {businessDistrict} {welfareList}')
    wb.save(f'D:/Download/export_data/boss/{keywords}.xlsx')

    driver.close()


if __name__ == '__main__':
    boss_spider('开发')
